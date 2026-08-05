"""
출장경비 계산 결과 -> Excel / PDF / 출장신청서 PDF 변환 유틸리티.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app_config import get_fuel_price_label, get_fuel_price_unit, get_fuel_type_label
from services.calculator import TripCalculationResult, get_allowance_reason, get_vehicle_type_label

_KOREAN_FONT_CANDIDATES = [
    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "C:/Windows/Fonts/malgun.ttf",
]
_registered_font_name: str | None = None


def format_won(amount: int | float, suffix: str = "원") -> str:
    """금액을 #,###원 형식으로 포맷한다."""
    if isinstance(amount, float):
        if amount == int(amount):
            amount = int(amount)
        else:
            return f"{amount:,.2f}{suffix}"
    return f"{int(amount):,}{suffix}"


def format_distance(km: float) -> str:
    return f"{km:,.2f} km"


def format_duration(minutes: float) -> str:
    total_min = int(round(minutes))
    hours, mins = divmod(total_min, 60)
    return f"{hours}시간 {mins}분" if hours > 0 else f"{mins}분"


def format_fuel_efficiency(km_per_unit: float, fuel_type: str = "gasoline") -> str:
    if fuel_type == "electric":
        return f"{km_per_unit:,.1f} km/kWh"
    return f"{km_per_unit:,.0f} km/L"


def format_liters(liters: float, fuel_type: str = "gasoline") -> str:
    if fuel_type == "electric":
        return f"{liters:,.2f} kWh"
    return f"{liters:,.2f} L"


def get_fuel_cost_label(fuel_type: str) -> str:
    return "충전비" if fuel_type == "electric" else "유류비"


def get_fuel_usage_label(fuel_type: str) -> str:
    return "사용 전력량" if fuel_type == "electric" else "사용 연료량"


def _register_korean_font() -> str:
    """PDF 한글 출력을 위해 시스템 폰트를 등록한다. (컨테이너: fonts-nanum 패키지)"""
    global _registered_font_name
    if _registered_font_name:
        return _registered_font_name

    font_name = "KoreanFont"
    for path in _KOREAN_FONT_CANDIDATES:
        try:
            pdfmetrics.registerFont(TTFont(font_name, path))
            _registered_font_name = font_name
            return font_name
        except Exception:
            continue

    _registered_font_name = "Helvetica"
    return _registered_font_name


def result_to_summary_dict(result: TripCalculationResult) -> dict[str, str]:
    """계산 결과를 화면/문서용 요약 딕셔너리로 변환한다."""
    route_lines = [result.departure, *result.destinations]

    return {
        "출장일자": result.trip_date.strftime("%Y-%m-%d"),
        "출발지": result.departure,
        "출장지": " -> ".join(result.destinations),
        "이동 경로": " -> ".join(route_lines),
        "차량 구분": get_vehicle_type_label(result.vehicle_type),
        "유종": get_fuel_type_label(result.fuel_type),
        "총 이동거리": format_distance(result.total_distance_km),
        "편도 거리": format_distance(result.one_way_distance_km),
        "예상 소요시간": format_duration(result.total_duration_min),
        get_fuel_price_label(result.fuel_type): format_won(result.fuel_price) + f"/{get_fuel_price_unit(result.fuel_type)}",
        "유가 출처": result.fuel_price_source,
        "연비": format_fuel_efficiency(result.fuel_efficiency, result.fuel_type),
        get_fuel_usage_label(result.fuel_type): format_liters(result.fuel_used_liters, result.fuel_type),
        get_fuel_cost_label(result.fuel_type): format_won(result.fuel_cost),
        "일비": format_won(result.daily_allowance),
        "일비 사유": get_allowance_reason(result.one_way_distance_km, result.destination_count),
        "최종 지급금액": format_won(result.total_payment),
    }


def export_to_excel(result: TripCalculationResult) -> bytes:
    """계산 결과를 Excel 파일(bytes)로 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "출장경비"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    summary = result_to_summary_dict(result)

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "출장경비 자동 산출 결과"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    row = 3
    for label, value in summary.items():
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="구간별 이동 상세").font = Font(bold=True, size=12)
    row += 1

    headers = ["구간", "출발", "도착", "거리(km)", "소요(분)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    row += 1

    for segment in result.route_segments:
        ws.cell(row=row, column=1, value=segment.get("segment_no", ""))
        ws.cell(row=row, column=2, value=segment.get("start", ""))
        ws.cell(row=row, column=3, value=segment.get("goal", ""))
        ws.cell(row=row, column=4, value=segment.get("distance_km", 0))
        ws.cell(row=row, column=5, value=segment.get("duration_min", 0))
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_to_pdf(result: TripCalculationResult) -> bytes:
    """계산 결과를 PDF 파일(bytes)로 생성한다."""
    font_name = _register_korean_font()
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20 * mm, leftMargin=20 * mm, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontName=font_name, fontSize=16, alignment=TA_CENTER, spaceAfter=12)

    summary = result_to_summary_dict(result)
    elements: list[Any] = [Paragraph("출장경비 자동 산출 결과", title_style), Spacer(1, 6 * mm)]

    table_data = [["항목", "내용"]]
    for label, value in summary.items():
        table_data.append([label, value])

    table = Table(table_data, colWidths=[45 * mm, 115 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_application_form(result: TripCalculationResult, applicant_name: str = "") -> bytes:
    """출장 신청서 양식 PDF를 생성한다."""
    font_name = _register_korean_font()
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20 * mm, leftMargin=20 * mm, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("FormTitle", parent=styles["Heading1"], fontName=font_name, fontSize=18, alignment=TA_CENTER, spaceAfter=16)
    subtitle_style = ParagraphStyle("FormSub", parent=styles["Normal"], fontName=font_name, fontSize=10, alignment=TA_CENTER, textColor=colors.grey)

    destination_text = ", ".join(result.destinations)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    form_rows = [
        ["출장일자", result.trip_date.strftime("%Y년 %m월 %d일")],
        ["신청자", applicant_name or "(입력 필요)"],
        ["출발지", result.departure],
        ["출장지", destination_text],
        ["유종", get_fuel_type_label(result.fuel_type)],
        ["차량 구분", get_vehicle_type_label(result.vehicle_type)],
        ["총 이동거리", format_distance(result.total_distance_km)],
        ["예상 소요시간", format_duration(result.total_duration_min)],
        [get_fuel_cost_label(result.fuel_type), format_won(result.fuel_cost)],
        ["일비", format_won(result.daily_allowance)],
        ["합계 지급금액", format_won(result.total_payment)],
        ["출장 목적", "(입력 필요)"],
        ["비고", get_allowance_reason(result.one_way_distance_km, result.destination_count)],
    ]

    elements: list[Any] = [
        Paragraph("출 장 신 청 서", title_style),
        Paragraph(f"작성일시: {generated_at}", subtitle_style),
        Spacer(1, 8 * mm),
    ]

    table = Table(form_rows, colWidths=[40 * mm, 120 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8EEF4")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15 * mm))

    sign_rows = [["부서장", "서명: _________________"], ["경영지원", "서명: _________________"]]
    sign_table = Table(sign_rows, colWidths=[40 * mm, 120 * mm])
    sign_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 20),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
    ]))
    elements.append(sign_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
